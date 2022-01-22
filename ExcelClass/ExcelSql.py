import openpyxl

from ExcelClass.BaseFilter import BaseFilter
from ExcelClass.ColSheet import ColSheet
from valueClass.ComplexValue import ComplexValue
from valueClass.equality_type import EqualityType


class ExcelSql:
    requestMatching = {
        "update": {
            "baseString": 'update {}',
            "required": ['set', 'where'],
        },
        "select": {
            "baseString": 'select * from {}',
            "required": ['where'],
        },
        "delete": {
            "baseString": 'delete from {}',
            "required": ['where'],
        }
    }
    workBook = None
    workSheets = None
    table: str
    filters: dict[str, list[BaseFilter]] = {}
    type: str
    fromSheetName: str
    currentWorkSheet = None
    datas = {}
    sheetToLoad: dict[str, list[ColSheet]] = {}
    offset: int = 2

    def __init__(self, xlsxName):
        self.workBook = openpyxl.load_workbook(filename=xlsxName)
        self.workSheets = {sheetName: self.workBook.get_sheet_by_name(sheetName) for sheetName in
                           self.workBook.sheetnames}

    def initReqSql(self, sqlType, table, fromSheetName, offset=2):
        self.fromSheetName = fromSheetName
        self.table = table
        self.type = sqlType
        self.offset = offset
        return self

    def addFilter(self, filter: BaseFilter):
        if filter.type not in self.filters.keys():
            self.filters[filter.type] = []
        self.filters[filter.type].append(filter)

    def generateSQL(self):
        self.loadAll()
        config = self.requestMatching[self.type]
        sqlReq = config['baseString'].format(self.table)
        for idx, row in self.datas[self.fromSheetName].items():
            newReq = sqlReq
            badReq = False
            for elem in config['required']:
                if badReq:
                    continue
                filters = self.getRowDatas(self.filters[elem], idx)
                res = self.makeStrFromParams(filters)
                if res:
                    newReq += f" {elem} {res}"
                else:
                    badReq = True
            if not badReq:
                print(f"{newReq};")

    def loadAll(self):
        allToLoad = []
        for filterType, filters in self.filters.items():
            for filter_ in filters:
                for v in filter_.value.getAllValue():
                    if v not in allToLoad:
                        allToLoad.append(v)

        for toLoad in allToLoad:
            if toLoad.sheetName not in self.sheetToLoad:
                self.sheetToLoad[toLoad.sheetName] = []
            self.sheetToLoad[toLoad.sheetName].append(toLoad)
        for sheetName, toLoads in self.sheetToLoad.items():
            dataSheet = {}
            for row in self.workSheets[sheetName].iter_rows(min_row=self.offset,
                                                            max_row=self.workSheets[sheetName].max_row):
                tmp = {}
                idx = str(row[0].row)
                for toLoad in toLoads:
                    tmp[toLoad.col] = self.workSheets[sheetName][f"{toLoad.col}{idx}"].value
                dataSheet[idx] = tmp
            self.datas[sheetName] = dataSheet

    def makeStrFromParams(self, filters: list[BaseFilter]):
        res = []
        isFirst = True
        for idx, filter in enumerate(filters):
            if filter.type == 'where':
                if isFirst:
                    isFirst = False
                    tmp = ""
                else:
                    tmp = "or " if filter.isOr else "and "
            else:
                tmp = ""
            tmp += f"`{filter.field}` {filter.equality} "
            val = filter.value.getVal()
            if type(val) is int:
                tmp += str(val)
            elif val is None:
                if filter.isRequired:
                    return False
                tmp += 'null'
            else:
                tmp += f"'{val}'"

            res.append(tmp)
        return " ".join(res) if filter.type == 'where' else ", ".join(res)

    def getRowDatas(self, filters: list[BaseFilter], idx):
        for filter in filters:
            filter.value.setRealVal(
                self.getValFromComplexValue(filter.value, idx) if filter.value.isComplex() else self.getValFromColSheet(
                    filter.value.value, idx))
        return filters

    def getValFromColSheet(self, colSheet, idx):
        return self.datas[colSheet.sheetName][idx][colSheet.col]

    def getValFromComplexValue(self, complexValue: ComplexValue, idx):

        from_Val = self.getValFromColSheet(complexValue.from_, idx)
        if complexValue.to_.sheetName == complexValue.value.sheetName:
            sheetName = complexValue.to_.sheetName
            for idx, row in self.datas[sheetName].items():
                fromParsed = from_Val.lower().strip()
                toParsed = row[complexValue.to_.col].lower().strip()
                if complexValue.equality == EqualityType.CONTAIN:
                    match = fromParsed in toParsed or toParsed in fromParsed
                elif complexValue.equality == EqualityType.EQUAL:
                    match = fromParsed == toParsed
                else:
                    match = False
                if match:
                    return row[complexValue.value.col]
        return None
